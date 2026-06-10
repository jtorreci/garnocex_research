#!/usr/bin/env python3
# =============================================================================
# TRANSPARENCY NOTE (public reproducibility package)
# =============================================================================
# This script is included for transparency only.  The full Scenario 2 pipeline
# requires the 23 confidential PT_*.xlsx operational files (raw commercial
# tonnages) which are NOT distributed with this package.  Those files contain
# commercially sensitive origin-destination flow data and are withheld
# consistent with the data availability statement in the manuscript.
#
# To reproduce the headline Scenario 2 indicators from public data, run:
#   python analysis/reproduce_scenarios.py
# which uses the anonymised file data/s2_observed_flows_public.csv (volumes
# rescaled to theoretical production; no commercial tonnages).
# =============================================================================
"""
compute_three_scenarios.py
===========================
Compute the 13-indicator framework for THREE scenarios and produce a
comparative table.

Scenario 1 – "Administrative" (A01 baseline)
    Each municipality is assigned to exactly ONE plant by network distance.
    Data from A01.Voronoi.

Scenario 2 – "Real flows" (regional operational data)
    Parse 23 PT_*.xlsx files containing REAL origin-destination flows
    (municipalities may send waste to MULTIPLE plants).
    Voronoi baseline comes from the THEORETICAL section of each file.
    NOTE: covers only 23 of 46 plants → indicators computed on the subset
    of municipalities that appear in the real data.

Scenario 3 – "Optimised" (A05 cost-optimal network)
    24 active plants from the A05 iterative cost-optimal algorithm.
    Data from optimal_assignments.csv.
"""

import os
import re
import sys
import warnings
from pathlib import Path

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore", category=UserWarning)

# ═══════════════════════════════════════════════════════════════════════
# PATHS
# ═══════════════════════════════════════════════════════════════════════
BASE = Path(__file__).resolve().parents[2]

# Scenario 1
A01_MUN_CSV = BASE / "A01.Voronoi/scripts/datos_red_real/datos_red_real_municipios.csv"

# Scenario 2
ANALYSIS_DIR_CANDIDATES = [
    BASE / "A03. Indicadores" / "Articulo_latex" / "Codigo" / "Análisis de Datos",
    BASE / "A03. Indicadores" / "Nueva carpeta" / "contenido" / "Articulo 1 Waste management" / "Análisis de Datos",
]

ANALYSIS_DIR = None
for candidate in ANALYSIS_DIR_CANDIDATES:
    if candidate.exists():
        ANALYSIS_DIR = candidate
        break

if ANALYSIS_DIR is None:
    raise FileNotFoundError("Cannot locate the 'Análisis de Datos' directory")

# Scenario 3
A05_CSV = BASE / "A05.Network_Optimization/outputs/optimal_assignments.csv"

# Output
OUT_DIR = BASE / "A03. Indicadores/scripts"
OUT_CSV = OUT_DIR / "three_scenarios_comparison.csv"
OUT_PLANT_S1 = OUT_DIR / "plant_level_s1_full.csv"
OUT_PLANT_S2 = OUT_DIR / "plant_level_s2_full.csv"
OUT_PLANT_S3 = OUT_DIR / "plant_level_s3_full.csv"

# ═══════════════════════════════════════════════════════════════════════
# PARAMETERS (same as A01)
# ═══════════════════════════════════════════════════════════════════════
C0   = 40_000    # EUR, minimum annual fixed cost
T0   = 5_000     # t/year, reference throughput for log cost
RHO  = 0.35      # EUR/t-km, transport tariff
V    = 0.35      # EUR/t, variable treatment cost
DOT  = 0.5       # t/hab/year, CDW endowment rate

# Consolidation map (Voronoi cell → active plant)
CONSOLIDATION = {
    2: 1, 12: 1,
    17: 3,
    15: 6,
    40: 19,
    32: 20,
    23: 22, 26: 22, 28: 22,
    34: 27, 35: 27, 37: 27,
    38: 36, 43: 36,
}


# ═══════════════════════════════════════════════════════════════════════
# NAME NORMALISATION
# ═══════════════════════════════════════════════════════════════════════
def norm(s):
    """Normalise a municipality name for matching across data sources."""
    if pd.isna(s):
        return ""
    s = str(s).strip().replace("\n", " ")
    s = s.encode("ascii", errors="ignore").decode("ascii")
    s = re.sub(r"^(La|Los|Las|El|Les)\s+", "", s, flags=re.IGNORECASE)
    s = re.sub(r"[^a-z0-9 ]", "", s.lower())
    return re.sub(r"\s+", " ", s).strip()


def consolidate(pid):
    """Map physical plant id to consolidated group id."""
    return CONSOLIDATION.get(int(pid), int(pid))


# ═══════════════════════════════════════════════════════════════════════
# COST MODEL
# ═══════════════════════════════════════════════════════════════════════
def fixed_cost(T):
    """Annual fixed cost for a plant with throughput T (t/year)."""
    if T <= 0:
        return C0
    return max(C0, C0 * (np.log2(T / T0) + 1))


def unit_treatment_cost(T):
    """Unit treatment cost (EUR/t) for throughput T."""
    if T <= 0:
        return np.inf
    return fixed_cost(T) / T + V


def unit_total_cost(d_km, T):
    """Unit total cost (EUR/t) at distance d_km from a plant with throughput T."""
    return RHO * d_km + unit_treatment_cost(T)


# ═══════════════════════════════════════════════════════════════════════
# STATISTICAL HELPERS
# ═══════════════════════════════════════════════════════════════════════
def weighted_quantile(values, weights, q):
    """q-th weighted quantile (0 <= q <= 1)."""
    values = np.asarray(values, dtype=float)
    weights = np.asarray(weights, dtype=float)
    mask = ~(np.isnan(values) | np.isnan(weights))
    values, weights = values[mask], weights[mask]
    if len(values) == 0:
        return np.nan
    idx = np.argsort(values)
    values, weights = values[idx], weights[idx]
    cum_w = np.cumsum(weights)
    cum_w_norm = (cum_w - 0.5 * weights) / cum_w[-1]
    return float(np.interp(q, cum_w_norm, values))


def gini_coefficient(values):
    """Gini coefficient of an array."""
    values = np.asarray(values, dtype=float)
    values = values[~np.isnan(values)]
    if len(values) == 0 or np.sum(values) == 0:
        return np.nan
    n = len(values)
    s = np.sort(values)
    idx = np.arange(1, n + 1)
    return float((2 * np.sum(idx * s) - (n + 1) * np.sum(s)) / (n * np.sum(s)))


def cvar(values, weights, alpha=0.9):
    """Conditional Value at Risk at level alpha (production-weighted)."""
    values = np.asarray(values, dtype=float)
    weights = np.asarray(weights, dtype=float)
    mask = ~(np.isnan(values) | np.isnan(weights))
    values, weights = values[mask], weights[mask]
    if len(values) == 0:
        return np.nan
    idx = np.argsort(values)
    values, weights = values[idx], weights[idx]
    cum_w = np.cumsum(weights)
    total_w = cum_w[-1]
    var_idx = min(np.searchsorted(cum_w, alpha * total_w), len(values) - 1)
    var_val = values[var_idx]
    tail = values >= var_val
    tw = weights[tail]
    if np.sum(tw) == 0:
        return float(var_val)
    return float(np.average(values[tail], weights=tw))


# ═══════════════════════════════════════════════════════════════════════
# INDICATOR COMPUTATION ENGINE
# ═══════════════════════════════════════════════════════════════════════
def compute_indicators(df_flows, df_voronoi, scenario_name, multi_flow=False):
    """
    Compute all 13 indicators from flow-level data.

    Parameters
    ----------
    df_flows : DataFrame
        Columns: mun_name, mun_norm, plant_id, prod_t, dist_km, cost_per_t
        If multi_flow=True, a municipality may appear in multiple rows (one per
        destination plant), each with a partial prod_t.
    df_voronoi : DataFrame
        Columns: mun_norm, plant_id_voronoi
        Defines the Voronoi (theoretical) assignment for the same universe of
        municipalities.  For Scenario 2 this comes from the THEORETICAL sections
        of the Excel files.
    scenario_name : str
    multi_flow : bool
        If True, municipalities can send waste to multiple plants.
    """
    print(f"\n{'='*70}")
    print(f"  SCENARIO: {scenario_name}")
    print(f"{'='*70}")

    # ---- Plant throughput ----
    plant_throughput = df_flows.groupby("plant_id")["prod_t"].sum()
    all_plants = sorted(plant_throughput.index)

    # ---- Unit treatment cost per plant ----
    plant_treat = plant_throughput.apply(unit_treatment_cost)

    # ---- Cost per flow row (per tonne) ----
    df = df_flows.copy()
    if "cost_per_t" not in df.columns or df["cost_per_t"].isna().all():
        df = df.merge(
            plant_treat.rename("C_treat"),
            left_on="plant_id", right_index=True, how="left",
        )
        df["cost_per_t"] = RHO * df["dist_km"] + df["C_treat"]
    else:
        df = df.merge(
            plant_treat.rename("C_treat"),
            left_on="plant_id", right_index=True, how="left",
        )

    df["tkm"] = df["prod_t"] * df["dist_km"]

    total_prod = df["prod_t"].sum()
    total_tkm = df["tkm"].sum()
    n_plants = len(all_plants)
    n_munis = df["mun_norm"].nunique()
    print(f"  Municipalities: {n_munis}")
    print(f"  Active plants : {n_plants}")
    print(f"  Total prod    : {total_prod:,.0f} t/yr")
    print(f"  Total t-km    : {total_tkm:,.0f}")

    # ---- Voronoi baseline: build theoretical tkm ----
    # For each municipality, the Voronoi baseline gives ONE plant assignment.
    # We need theoretical t-km: sum over all production × (theoretical distance
    # to Voronoi plant).
    # Since we may not have theoretical distances for all combos, we approximate:
    # For scenario 1 & 3, the theoretical baseline is the nearest-plant
    # assignment by road distance from A01.
    #   -> use the road-distance baseline stored in the A01 municipality file.
    # For scenario 2, the theoretical data from the Excel files already contains
    #   the theoretical t-km.

    # Build Voronoi lookup: mun_norm -> plant_id_voronoi
    vor_map = dict(zip(df_voronoi["mun_norm"], df_voronoi["plant_id_voronoi"]))

    # For theoretical tkm, we need to know the theoretical distance for each
    # municipality to its Voronoi plant.  If available in df_voronoi, use it.
    if "theo_dist_km" in df_voronoi.columns:
        vor_dist_map = dict(zip(df_voronoi["mun_norm"], df_voronoi["theo_dist_km"]))
    else:
        vor_dist_map = {}

    if "theo_tkm" in df_voronoi.columns:
        vor_tkm_map = dict(zip(df_voronoi["mun_norm"], df_voronoi["theo_tkm"]))
    else:
        vor_tkm_map = {}

    if "theo_prod_t" in df_voronoi.columns:
        vor_prod_map = dict(zip(df_voronoi["mun_norm"], df_voronoi["theo_prod_t"]))
    else:
        vor_prod_map = {}

    # Assign Voronoi plant to each flow row
    df["plant_voronoi"] = df["mun_norm"].map(vor_map)
    df["is_tp"] = (df["plant_id"] == df["plant_voronoi"]).astype(int)

    # ---- Theoretical t-km ----
    # For each municipality, theoretical tkm = theoretical prod × theoretical dist
    # Aggregate by Voronoi plant
    # We need this at plant level and system level

    # Build theoretical flow DataFrame (one row per mun in Voronoi assignment)
    theo_rows = []
    for _, vrow in df_voronoi.iterrows():
        mn = vrow["mun_norm"]
        pid_v = vrow["plant_id_voronoi"]
        # Get production from real flows for this municipality
        # (sum across all destination plants if multi-flow)
        mun_real = df[df["mun_norm"] == mn]
        if len(mun_real) == 0:
            continue
        real_total_prod = mun_real["prod_t"].sum()

        # Theoretical production: use the value from Voronoi data if available
        if mn in vor_prod_map and not pd.isna(vor_prod_map[mn]):
            theo_prod = vor_prod_map[mn]
        else:
            theo_prod = real_total_prod

        # Theoretical distance
        if mn in vor_dist_map and not pd.isna(vor_dist_map[mn]):
            theo_d = vor_dist_map[mn]
        else:
            # Fallback: use the distance from real data if the mun IS assigned to
            # the Voronoi plant in the real data
            match = mun_real[mun_real["plant_id"] == pid_v]
            if len(match) > 0:
                theo_d = match["dist_km"].iloc[0]
            else:
                theo_d = np.nan

        # Theoretical tkm
        if mn in vor_tkm_map and not pd.isna(vor_tkm_map[mn]):
            theo_tkm = vor_tkm_map[mn]
        else:
            theo_tkm = theo_prod * theo_d if not pd.isna(theo_d) else np.nan

        theo_rows.append({
            "mun_norm": mn,
            "plant_id_voronoi": pid_v,
            "theo_prod": theo_prod,
            "theo_dist": theo_d,
            "theo_tkm": theo_tkm,
        })

    df_theo = pd.DataFrame(theo_rows)

    total_theo_tkm = df_theo["theo_tkm"].sum()
    print(f"  Total theo t-km: {total_theo_tkm:,.0f}")

    # ---- TP/FP/FN per plant ----
    # In multi-flow, a municipality can contribute partial TP to one plant
    # and partial FP to another.
    # TP for plant i: production sent to plant i from municipalities whose
    #                 Voronoi assignment is ALSO plant i.
    # FP for plant i: production sent to plant i from municipalities whose
    #                 Voronoi assignment is a DIFFERENT plant.
    # FN for plant i: production that Voronoi says should go to plant i, but
    #                 the municipality sends it elsewhere (partially or fully).

    tp_by_plant = df[df["is_tp"] == 1].groupby("plant_id")["prod_t"].sum()
    fp_by_plant = df[df["is_tp"] == 0].groupby("plant_id")["prod_t"].sum()

    # FN: for each Voronoi plant, sum prod from municipalities in its cell that
    # sent waste to OTHER plants
    fn_by_plant = df[df["is_tp"] == 0].groupby("plant_voronoi")["prod_t"].sum()

    # ---- Plant-level DataFrame ----
    plant_df = pd.DataFrame(index=all_plants)
    plant_df.index.name = "plant_id"
    plant_df["T_i"] = plant_throughput.reindex(all_plants).fillna(0)
    plant_df["TP"] = tp_by_plant.reindex(all_plants).fillna(0)
    plant_df["FP"] = fp_by_plant.reindex(all_plants).fillna(0)
    plant_df["FN"] = fn_by_plant.reindex(all_plants).fillna(0)

    # Precision, Recall
    plant_df["Prec_i"] = plant_df["TP"] / (plant_df["TP"] + plant_df["FP"])
    plant_df["Rec_i"] = plant_df["TP"] / (plant_df["TP"] + plant_df["FN"])
    plant_df["Prec_i"] = plant_df["Prec_i"].fillna(0)
    plant_df["Rec_i"] = plant_df["Rec_i"].fillna(0)

    # ---- 1. C_i: unit total cost per plant ----
    agg = df.groupby("plant_id").agg(
        total_cost_x_prod=pd.NamedAgg("cost_x_prod", "sum"),
        total_tkm=pd.NamedAgg("tkm", "sum"),
        total_prod=pd.NamedAgg("prod_t", "sum"),
    ) if "cost_x_prod" in df.columns else None

    # Simpler: production-weighted average cost per plant
    df["cost_x_prod"] = df["cost_per_t"] * df["prod_t"]
    agg = df.groupby("plant_id").agg(
        total_cost_x_prod=("cost_x_prod", "sum"),
        total_tkm=("tkm", "sum"),
        total_prod=("prod_t", "sum"),
    )
    plant_df["C_i"] = agg["total_cost_x_prod"] / agg["total_prod"]

    # ---- 2. U_i: capacity utilisation ----
    plant_df["U_i"] = np.nan  # No capacity data available

    # ---- 3. D_i^(90): 90th-percentile weighted accessibility (km) ----
    d90_vals = {}
    for pid in all_plants:
        sub = df[df["plant_id"] == pid]
        d90_vals[pid] = weighted_quantile(
            sub["dist_km"].values, sub["prod_t"].values, 0.9
        )
    plant_df["D90_i"] = pd.Series(d90_vals)

    # ---- 6. Leak_dg_i: cost-penalised leakage ----
    leak_vals = {}
    for pid in all_plants:
        # Municipalities whose Voronoi assignment is this plant
        munis_in_voronoi = df_voronoi[df_voronoi["plant_id_voronoi"] == pid]["mun_norm"].unique()
        if len(munis_in_voronoi) == 0:
            leak_vals[pid] = 0.0
            continue

        # Get the total Voronoi production for this cell
        voronoi_total_prod = df[df["mun_norm"].isin(munis_in_voronoi)]["prod_t"].sum()
        if voronoi_total_prod == 0:
            leak_vals[pid] = 0.0
            continue

        # Find flows from Voronoi-cell municipalities that go to OTHER plants
        cell_flows = df[(df["mun_norm"].isin(munis_in_voronoi)) & (df["plant_id"] != pid)]
        if len(cell_flows) == 0:
            leak_vals[pid] = 0.0
            continue

        # Cost at the Voronoi plant for each leaked municipality
        T_pid = plant_df.loc[pid, "T_i"]
        c_treat_vor = unit_treatment_cost(T_pid) if T_pid > 0 else np.inf

        penalty_sum = 0.0
        for _, frow in cell_flows.iterrows():
            # Voronoi cost: transport to this plant + treatment at this plant
            d_vor = vor_dist_map.get(frow["mun_norm"], np.nan)
            if pd.isna(d_vor):
                # If we don't know the theoretical distance, skip this flow
                continue
            g_voronoi = RHO * d_vor + c_treat_vor
            g_real = frow["cost_per_t"]
            delta_g = max(0.0, g_real - g_voronoi)
            penalty_sum += frow["prod_t"] * delta_g

        leak_vals[pid] = penalty_sum / voronoi_total_prod

    plant_df["Leak_dg_i"] = pd.Series(leak_vals)

    # ---- 7. IET_i: transport efficiency per plant ----
    real_tkm_by_plant = df.groupby("plant_id")["tkm"].sum()
    theo_tkm_by_plant = df_theo.groupby("plant_id_voronoi")["theo_tkm"].sum()
    plant_df["tkm_real"] = real_tkm_by_plant.reindex(all_plants).fillna(0)
    plant_df["tkm_theo"] = theo_tkm_by_plant.reindex(all_plants).fillna(0)
    plant_df["IET_i"] = np.where(
        plant_df["tkm_real"] > 0,
        100.0 * plant_df["tkm_theo"] / plant_df["tkm_real"],
        np.nan,
    )

    # ═══════════════════════════════════════════════════════════════════
    # NETWORK-LEVEL INDICATORS
    # ═══════════════════════════════════════════════════════════════════

    # 1. C_bar: production-weighted average unit cost
    C_bar = np.average(df["cost_per_t"], weights=df["prod_t"])

    # 2. Gini(C_m) – use per-municipality cost
    # In multi-flow, a municipality's cost is the weighted average across
    # its destination plants
    mun_cost = df.groupby("mun_norm").apply(
        lambda g: np.average(g["cost_per_t"], weights=g["prod_t"])
    )
    Gini_val = gini_coefficient(mun_cost.values)

    # 3. MAD_C: production-weighted mean absolute deviation
    mun_prod = df.groupby("mun_norm")["prod_t"].sum()
    mad_values = np.abs(mun_cost.values - C_bar)
    MAD_C = np.average(mad_values, weights=mun_prod.reindex(mun_cost.index).values)

    # Export per-municipality cost components (transport / treatment / total) for KDE figures.
    # transport = RHO*d (clean); treatment = total - transport (consistent with each scenario's total).
    _kde_tag = {"1. Administrative (A01) - full network": "s1",
                "2full. Real flows with conservative full-network closure": "s2",
                "3. Optimised (A05) - full network": "s3"}.get(scenario_name)
    if _kde_tag:
        muni_transport = df.groupby("mun_norm").apply(
            lambda gg: np.average(RHO * gg["dist_km"], weights=gg["prod_t"]))
        kde_df = pd.DataFrame({
            "mun": mun_cost.index,
            "transport": muni_transport.reindex(mun_cost.index).values,
            "total": mun_cost.values,
            "prod": mun_prod.reindex(mun_cost.index).values,
        })
        kde_df["treatment"] = kde_df["total"] - kde_df["transport"]
        kde_df.to_csv(OUT_DIR / f"muni_costs_{_kde_tag}.csv", index=False)
        print(f"  [KDE export] muni_costs_{_kde_tag}.csv ({len(kde_df)} munis)")

    # 4. CVaR_0.9(D): on distances, production-weighted
    CVaR_D = cvar(df["dist_km"].values, df["prod_t"].values, alpha=0.9)

    # 5. micro-F1
    total_TP = plant_df["TP"].sum()
    total_FP = plant_df["FP"].sum()
    total_FN = plant_df["FN"].sum()
    denom = 2 * total_TP + total_FP + total_FN
    micro_F1 = 2 * total_TP / denom if denom > 0 else 0.0

    # 6. IGET: global transport efficiency
    total_real_tkm = df["tkm"].sum()
    total_theo_tkm_val = df_theo["theo_tkm"].sum()
    IGET = 100.0 * total_theo_tkm_val / total_real_tkm if total_real_tkm > 0 else np.nan

    # 7. CV(T_i)
    cv_ti = (plant_df["T_i"].std() / plant_df["T_i"].mean()
             if plant_df["T_i"].mean() > 0 else np.nan)

    # ---- Additional ----
    total_system_cost = (df["cost_per_t"] * df["prod_t"]).sum()

    # Irrational leakage cost
    irrational_cost = 0.0
    for pid in all_plants:
        munis_in_voronoi = set(
            df_voronoi[df_voronoi["plant_id_voronoi"] == pid]["mun_norm"]
        )
        cell_flows = df[(df["mun_norm"].isin(munis_in_voronoi)) & (df["plant_id"] != pid)]
        if len(cell_flows) == 0:
            continue
        T_pid = plant_df.loc[pid, "T_i"]
        c_treat_vor = unit_treatment_cost(T_pid) if T_pid > 0 else np.inf
        for _, frow in cell_flows.iterrows():
            d_vor = vor_dist_map.get(frow["mun_norm"], np.nan)
            if pd.isna(d_vor):
                continue
            g_voronoi = RHO * d_vor + c_treat_vor
            g_real = frow["cost_per_t"]
            if g_real > g_voronoi:
                irrational_cost += frow["prod_t"] * (g_real - g_voronoi)

    # Redundancy check
    IGCR = 10000.0 / IGET if IGET and IGET > 0 else np.nan
    IGSD = 10000.0 / IGET - 100.0 if IGET and IGET > 0 else np.nan

    # ---- Print ----
    print(f"\n  --- NETWORK-LEVEL INDICATORS ---")
    print(f"  C_bar              : {C_bar:.2f} EUR/t")
    print(f"  Gini(C_m)          : {Gini_val:.4f}")
    print(f"  MAD_C              : {MAD_C:.2f} EUR/t")
    print(f"  CVaR_0.9(D)        : {CVaR_D:.1f} km")
    print(f"  micro-F1           : {micro_F1:.4f}")
    print(f"  IGET               : {IGET:.1f}%")
    print(f"  CV(T_i)            : {cv_ti:.3f}")
    print(f"  Total system cost  : {total_system_cost:,.0f} EUR/yr")
    print(f"  Total t-km         : {total_real_tkm:,.0f}")
    print(f"  Irrational leak $  : {irrational_cost:,.0f} EUR/yr")
    print(f"  IGCR               : {IGCR:.2f}")
    print(f"  IGSD               : {IGSD:.2f}")
    print(f"  Verify IGCR=10000/IGET : {10000.0/IGET:.2f}" if IGET and IGET > 0 else "")
    print(f"  Verify IGSD=IGCR-100   : {IGCR-100:.2f}" if not np.isnan(IGCR) else "")

    # ---- Plant-level summary ----
    cols_show = ["T_i", "C_i", "D90_i", "Prec_i", "Rec_i", "Leak_dg_i", "IET_i"]
    print(f"\n  --- PLANT-LEVEL INDICATORS (sample) ---")
    print(plant_df[cols_show].head(10).to_string(float_format=lambda x: f"{x:.2f}"))

    return {
        "scenario": scenario_name,
        # Network-level
        "C_bar": C_bar,
        "Gini": Gini_val,
        "MAD_C": MAD_C,
        "CVaR_0.9(D)": CVaR_D,
        "micro-F1": micro_F1,
        "IGET": IGET,
        "CV(T_i)": cv_ti,
        # Additional
        "Total_system_cost_EUR": total_system_cost,
        "Total_tkm": total_real_tkm,
        "Irrational_leak_cost_EUR": irrational_cost,
        "IGCR": IGCR,
        "IGSD": IGSD,
        "N_plants": n_plants,
        "N_municipalities": n_munis,
        "Total_prod_t": total_prod,
        # For plant-level averages
        "mean_C_i": plant_df["C_i"].mean(),
        "mean_D90_i": plant_df["D90_i"].mean(),
        "mean_Prec_i": plant_df["Prec_i"].mean(),
        "mean_Rec_i": plant_df["Rec_i"].mean(),
        "mean_Leak_dg_i": plant_df["Leak_dg_i"].mean(),
        "mean_IET_i": plant_df["IET_i"].mean(),
        "median_IET_i": plant_df["IET_i"].median(),
        # plant_df for further analysis
        "_plant_df": plant_df,
    }


# ═══════════════════════════════════════════════════════════════════════
# SCENARIO 1: Administrative (A01 baseline)
# ═══════════════════════════════════════════════════════════════════════
def load_scenario1():
    """Load data for Scenario 1 (Administrative / A01 baseline)."""
    print("\n[S1] Loading A01 data...")

    df_mun = pd.read_csv(A01_MUN_CSV)

    # Consolidate plant IDs
    df_mun["plant_id"] = df_mun["planta_id"].apply(consolidate)

    # Normalise names
    df_mun["mun_norm"] = df_mun["municipio"].apply(norm)

    # Build flow DataFrame
    df_flows = pd.DataFrame({
        "mun_name": df_mun["municipio"],
        "mun_norm": df_mun["mun_norm"],
        "plant_id": df_mun["plant_id"],
        "prod_t": df_mun["produccion_t"],
        "dist_km": df_mun["distancia_km"],
        "cost_per_t": df_mun["C_tot"],
    })

    # Build theoretical baseline DataFrame using nearest-plant assignment
    # by road distance from the A01 municipality table.
    df_voronoi = pd.DataFrame({
        "mun_norm": df_mun["mun_norm"],
        "plant_id_voronoi": df_mun["plant_id"],
        "theo_dist_km": df_mun["distancia_km"],
    }).drop_duplicates(subset="mun_norm", keep="first")

    print(f"  Loaded {len(df_flows)} municipal flows, {df_flows['plant_id'].nunique()} plants")
    return df_flows, df_voronoi


# ═══════════════════════════════════════════════════════════════════════
# SCENARIO 2: Real flows (regional operational Excel files)
# ═══════════════════════════════════════════════════════════════════════
# Mapping from Excel filename stem to a synthetic plant identifier.
# We use the plant municipality name from the Resumen/A01 data.
# Plants NOT in A01 get synthetic negative IDs.
EXCEL_PLANT_MAP = {
    # file_stem : (plant_id or synthetic, plant_city_name)
    "PT_Aldeanueva":              (-1, "Aldeanueva del Camino"),
    "PT_Badajoz":                 (18, "Badajoz"),
    "PT_Brozas":                  (44, "Brozas"),
    "PT_Cadalso":                 (31, "Cadalso"),
    "PT_Caceres":                 (46, "Caceres"),
    "PT_CabezueladelValle":       (22, "Cabezuela del Valle"),
    "PT_Ceclavin":                (24, "Ceclavin"),
    "PT_Coria":                   (-2, "Coria"),
    "PT_Don Benito":              ( 3, "Don Benito"),
    "PT_Jaraiz de la Vera":       (-3, "Jaraiz de la Vera"),
    "PT_Jarandilla de la Vera":   (-4, "Jarandilla de la Vera"),
    "PT_Logrosan":                (45, "Logrosan"),
    "PT_Montanchez":              (41, "Montanchez"),
    "PT_Montehermoso":            (25, "Montehermoso"),
    "PT_Moraleja":                (20, "Moraleja"),
    "PT_Navalmoral de la Mata":   (27, "Navalmoral de la Mata"),
    "PT_Plasencia":               (29, "Plasencia"),
    "PT_Serradilla":              (33, "Serradilla"),
    "PT_Trujillo":                (19, "Trujillo"),
    "PT_Valencia de Alcantara":   (39, "Valencia de Alcantara"),
}

# Handle files with encoding artefacts in their names
_STEM_ALIASES = {
    # If the actual filename on disk has garbled chars, map it here
}


def _find_stem(filename):
    """Match a filename (without extension) to an EXCEL_PLANT_MAP key."""
    stem = Path(filename).stem
    if stem in EXCEL_PLANT_MAP:
        return stem
    # Try ascii-normalised matching
    stem_n = norm(stem.replace("PT_", ""))
    for key in EXCEL_PLANT_MAP:
        key_n = norm(key.replace("PT_", ""))
        if stem_n == key_n:
            return key
    return None


def _parse_excel_file(filepath):
    """
    Parse one PT_*.xlsx file, returning (real_rows, theo_rows).
    Each list contains dicts with keys:
        mun_name, inhabitants, dist_m, dist_km, tonnes, tkm, cost_total, cost_avg
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb.active
    max_col = ws.max_column
    max_row = ws.max_row

    def read_data_row(r, col_offset=0):
        """Read a data row starting at column col_offset+1."""
        c0 = col_offset
        mun_name = ws.cell(r, c0 + 1).value
        if mun_name is None or (isinstance(mun_name, str) and
                                 mun_name.strip().upper() in ("", "TOTAL")):
            return None
        inhabitants = ws.cell(r, c0 + 2).value
        dist_m = ws.cell(r, c0 + 3).value
        dist_km = ws.cell(r, c0 + 4).value
        tonnes = ws.cell(r, c0 + 5).value
        tkm = ws.cell(r, c0 + 6).value
        cost_total = ws.cell(r, c0 + 7).value
        cost_avg = ws.cell(r, c0 + 8).value

        # Validate: skip header rows
        if isinstance(mun_name, str) and "Municipio" in mun_name:
            return None
        if isinstance(mun_name, str) and "DATOS" in mun_name.upper():
            return None

        # Skip rows where tonnes is not numeric
        try:
            tonnes = float(tonnes) if tonnes is not None else 0.0
        except (ValueError, TypeError):
            return None

        try:
            dist_km = float(dist_km) if dist_km is not None else 0.0
            dist_m = float(dist_m) if dist_m is not None else 0.0
        except (ValueError, TypeError):
            dist_km = 0.0
            dist_m = 0.0

        try:
            tkm = float(tkm) if tkm is not None else 0.0
        except (ValueError, TypeError):
            tkm = tonnes * dist_km

        try:
            cost_total = float(cost_total) if cost_total is not None else 0.0
        except (ValueError, TypeError):
            cost_total = 0.0

        try:
            cost_avg = float(cost_avg) if cost_avg is not None else 0.0
        except (ValueError, TypeError):
            cost_avg = 0.0

        return {
            "mun_name": str(mun_name).strip() if mun_name else "",
            "inhabitants": float(inhabitants) if inhabitants else 0.0,
            "dist_m": dist_m,
            "dist_km": dist_km,
            "tonnes": tonnes,
            "tkm": tkm,
            "cost_total": cost_total,
            "cost_avg": cost_avg,
        }

    real_rows = []
    theo_rows = []

    # Detect layout
    is_side_by_side = max_col >= 19

    if is_side_by_side:
        # Side-by-side: Real in cols 1-8, Theoretical in cols 10-17
        # Row 1: headers, Row 2: column labels, Row 3+: data
        for r in range(3, max_row + 1):
            v = ws.cell(r, 1).value
            if v is not None and isinstance(v, str) and "TOTAL" in v.upper():
                # After TOTAL, there may be more real rows but we stop
                break
            row_data = read_data_row(r, col_offset=0)
            if row_data is not None and row_data["tonnes"] > 0:
                real_rows.append(row_data)

        # Theoretical data (cols 10-17)
        for r in range(3, max_row + 1):
            v = ws.cell(r, 10).value
            if v is None:
                continue
            if isinstance(v, str) and "TOTAL" in v.upper():
                break
            if isinstance(v, str) and ("Municipio" in v or "DATOS" in v.upper()):
                continue
            row_data = read_data_row(r, col_offset=9)
            if row_data is not None and row_data["tonnes"] > 0:
                theo_rows.append(row_data)
    else:
        # Stacked layout: Real block, then blank rows, then "DATOS TEORICOS",
        # then header row, then theoretical data
        section = "real"
        for r in range(3, max_row + 1):
            v = ws.cell(r, 1).value
            if v is None:
                continue
            vs = str(v).strip().upper()
            # Normalise to ASCII for keyword detection (TEÓRICOS -> TERICOS)
            vs_ascii = vs.encode("ascii", errors="ignore").decode("ascii")
            if vs == "TOTAL":
                continue
            if "TERIC" in vs_ascii or "TEOR" in vs_ascii:
                section = "theo_header"
                continue
            if section == "theo_header" and "Municipio" in str(v):
                section = "theo"
                continue
            if section == "theo_header":
                # This might be the first data row of theoretical section
                # (some files go directly to data after DATOS TEORICOS header)
                section = "theo"

            row_data = read_data_row(r, col_offset=0)
            if row_data is None:
                continue
            if row_data["tonnes"] <= 0:
                continue

            if section == "real":
                real_rows.append(row_data)
            elif section == "theo":
                theo_rows.append(row_data)

    wb.close()
    return real_rows, theo_rows


def load_scenario2():
    """Load data for Scenario 2 (Real flows from Excel files)."""
    print("\n[S2] Loading regional operational data from Excel files...")

    all_real_flows = []
    all_theo_flows = []
    files_parsed = 0
    files_skipped = []

    for f in sorted(ANALYSIS_DIR.iterdir()):
        if not f.name.startswith("PT_") or not f.name.endswith(".xlsx"):
            continue

        stem_key = _find_stem(f.name)

        # Handle files not in the predefined map by trying partial match
        if stem_key is None:
            # Try to match by removing encoding artefacts
            stem_clean = f.stem.encode("ascii", errors="ignore").decode("ascii")
            stem_clean = re.sub(r"[^a-zA-Z0-9_ ]", "", stem_clean)
            for key in EXCEL_PLANT_MAP:
                key_clean = key.encode("ascii", errors="ignore").decode("ascii")
                key_clean = re.sub(r"[^a-zA-Z0-9_ ]", "", key_clean)
                if stem_clean == key_clean:
                    stem_key = key
                    break

        if stem_key is None:
            files_skipped.append(f.name)
            continue

        plant_id, plant_city = EXCEL_PLANT_MAP[stem_key]

        try:
            real_rows, theo_rows = _parse_excel_file(f)
        except Exception as e:
            print(f"  WARNING: Failed to parse {f.name}: {e}")
            files_skipped.append(f.name)
            continue

        for row in real_rows:
            all_real_flows.append({
                "plant_id": plant_id,
                "plant_city": plant_city,
                "mun_name": row["mun_name"],
                "mun_norm": norm(row["mun_name"]),
                "prod_t": row["tonnes"],
                "dist_km": row["dist_km"],
                "tkm": row["tkm"],
                "cost_total": row["cost_total"],
                "cost_avg": row["cost_avg"],
            })

        for row in theo_rows:
            all_theo_flows.append({
                "plant_id_voronoi": plant_id,
                "plant_city": plant_city,
                "mun_name": row["mun_name"],
                "mun_norm": norm(row["mun_name"]),
                "theo_prod_t": row["tonnes"],
                "theo_dist_km": row["dist_km"],
                "theo_tkm": row["tkm"],
            })

        files_parsed += 1
        print(f"  Parsed {f.name}: {len(real_rows)} real, {len(theo_rows)} theoretical")

    if files_skipped:
        print(f"  Skipped files (no map): {files_skipped}")
    print(f"  Total files parsed: {files_parsed}")

    # --- Also handle the 3 files with encoding artefacts in names ---
    # PT_Ca+¦averal -> Cañaveral, PT_Casta+¦ar -> Castañar, PT_Nu+¦omoral -> Nuñomoral
    _extra_map = {
        "caaveral": (30, "Canaveral"),
        "castaar de ibor": (42, "Castanar de Ibor"),
        "nuomoral": (21, "Nunomoral"),
    }
    for f in sorted(ANALYSIS_DIR.iterdir()):
        if not f.name.startswith("PT_") or not f.name.endswith(".xlsx"):
            continue
        stem_key = _find_stem(f.name)
        if stem_key is not None:
            continue  # already parsed
        # Try extra map
        stem_n = norm(f.stem.replace("PT_", ""))
        found = False
        for pattern, (pid, pname) in _extra_map.items():
            if pattern in stem_n or stem_n in pattern:
                found = True
                try:
                    real_rows, theo_rows = _parse_excel_file(f)
                except Exception as e:
                    print(f"  WARNING: Failed to parse {f.name}: {e}")
                    break
                for row in real_rows:
                    all_real_flows.append({
                        "plant_id": pid,
                        "plant_city": pname,
                        "mun_name": row["mun_name"],
                        "mun_norm": norm(row["mun_name"]),
                        "prod_t": row["tonnes"],
                        "dist_km": row["dist_km"],
                        "tkm": row["tkm"],
                        "cost_total": row["cost_total"],
                        "cost_avg": row["cost_avg"],
                    })
                for row in theo_rows:
                    all_theo_flows.append({
                        "plant_id_voronoi": pid,
                        "plant_city": pname,
                        "mun_name": row["mun_name"],
                        "mun_norm": norm(row["mun_name"]),
                        "theo_prod_t": row["tonnes"],
                        "theo_dist_km": row["dist_km"],
                        "theo_tkm": row["tkm"],
                    })
                files_parsed += 1
                print(f"  Parsed {f.name} (extra): {len(real_rows)} real, {len(theo_rows)} theoretical")
                break
        # end for pattern

    print(f"  Total files parsed (incl. extra): {files_parsed}")

    df_real = pd.DataFrame(all_real_flows)
    df_theo = pd.DataFrame(all_theo_flows)

    # Remove rows with zero production
    df_real = df_real[df_real["prod_t"] > 0].copy()
    df_theo = df_theo[df_theo["theo_prod_t"] > 0].copy()

    # Build flow DataFrame for the compute engine
    # Revalue Scenario 2 with the same transport+treatment model used in S1/S3.
    # The Excel monetary columns are retained only as source metadata; they are
    # not used for the comparative indicators because the transport unit there
    # is EUR/km rather than EUR/t-km.
    df_flows = pd.DataFrame({
        "mun_name": df_real["mun_name"],
        "mun_norm": df_real["mun_norm"],
        "plant_id": df_real["plant_id"],
        "prod_t": df_real["prod_t"],
        "dist_km": df_real["dist_km"],
        "cost_per_t": np.nan,
    })

    # Build Voronoi DataFrame
    # For Scenario 2, the Voronoi baseline is the THEORETICAL assignment from
    # the same Excel files.
    # For municipalities that appear in multiple theoretical files (unlikely),
    # keep the first occurrence.
    df_voronoi = df_theo[["mun_norm", "plant_id_voronoi", "theo_dist_km",
                           "theo_tkm", "theo_prod_t"]].copy()
    df_voronoi = df_voronoi.rename(columns={
        "theo_dist_km": "theo_dist_km",
        "theo_tkm": "theo_tkm",
        "theo_prod_t": "theo_prod_t",
    })
    # Drop duplicates (keep first per municipality)
    df_voronoi = df_voronoi.drop_duplicates(subset="mun_norm", keep="first")

    n_real_munis = df_flows["mun_norm"].nunique()
    n_real_plants = df_flows["plant_id"].nunique()
    n_theo_munis = df_voronoi["mun_norm"].nunique()
    print(f"\n  Real flows: {len(df_flows)} rows, {n_real_munis} municipalities, "
          f"{n_real_plants} plants")
    print(f"  Theoretical: {len(df_voronoi)} rows, {n_theo_munis} municipalities")
    print(f"  NOTE: These cover only {n_real_plants} of 46 plants (partial Extremadura).")
    print(f"  Indicators are computed on the subset of municipalities in the real data.")

    return df_flows, df_voronoi


# ═══════════════════════════════════════════════════════════════════════
# SCENARIO 3: Optimised (A05 cost-optimal network)
# ═══════════════════════════════════════════════════════════════════════
def load_scenario3():
    """Load data for Scenario 3 (A05 optimised network)."""
    print("\n[S3] Loading A05 optimal assignments...")

    df_a05 = pd.read_csv(A05_CSV)
    df_mun = pd.read_csv(A01_MUN_CSV)
    df_mun["mun_norm"] = df_mun["municipio"].apply(norm)
    df_mun["plant_id"] = df_mun["planta_id"].apply(consolidate)

    # Build flow DataFrame
    df_flows = pd.DataFrame({
        "mun_name": df_a05["name"],
        "mun_norm": df_a05["mun_key"],
        "plant_id": df_a05["plant_id"].astype(int),
        "prod_t": df_a05["prod"],
        "dist_km": df_a05["dist_km"],
        "cost_per_t": df_a05["cost"],
    })

    # Reuse the same road-distance theoretical baseline as Scenario 1.
    df_voronoi = pd.DataFrame({
        "mun_norm": df_mun["mun_norm"],
        "plant_id_voronoi": df_mun["plant_id"],
        "theo_dist_km": df_mun["distancia_km"],
    }).drop_duplicates(subset="mun_norm", keep="first")

    print(f"  Loaded {len(df_flows)} flows, {df_flows['plant_id'].nunique()} plants")
    return df_flows, df_voronoi


# ═══════════════════════════════════════════════════════════════════════
# SUPPORT CONSTRUCTIONS: AUDITED SUBSET AND CONSERVATIVE CLOSURE
# ═══════════════════════════════════════════════════════════════════════
def restrict_to_municipal_subset(df_flows, df_voronoi, mun_subset):
    """Restrict flows and baseline to a given municipality subset."""
    mun_subset = set(mun_subset)
    flows_sub = df_flows[df_flows["mun_norm"].isin(mun_subset)].copy()
    vor_sub = df_voronoi[df_voronoi["mun_norm"].isin(mun_subset)].copy()
    return flows_sub, vor_sub


def normalize_observed_flows_to_baseline(flows_obs, flows1_full):
    """
    Enforce municipal mass balance against the S1 theoretical production.

    For each municipality:
      - if observed flow total > baseline production, scale observed flows down
        proportionally so they sum exactly to the baseline production;
      - if observed flow total <= baseline production, keep observed flows as they are.
    """
    baseline_prod = (
        flows1_full.drop_duplicates(subset="mun_norm")
        .set_index("mun_norm")["prod_t"]
        .to_dict()
    )

    df = flows_obs.copy()
    obs_totals = df.groupby("mun_norm")["prod_t"].transform("sum")
    target = df["mun_norm"].map(baseline_prod)
    scale = np.where(
        target.notna() & (obs_totals > 0) & (obs_totals > target),
        target / obs_totals,
        1.0,
    )
    df["prod_t"] = df["prod_t"] * scale
    return df


def build_s2_conservative_closure(flows2_obs, flows1_full, vor1_full):
    """
    Complete the observed S2 subnetwork to the full municipality universe.

    For each municipality m in the full S1 baseline:
      residual_m = G_m(full baseline) - observed_flow_m_to_audited_plants

    The residual is conservatively assigned to the S1 proximity-baseline plant.
    This preserves observed audited flows while avoiding arbitrary imputation for
    unaudited plants.
    """
    # Baseline one-row-per-municipality reference
    s1_ref = flows1_full.drop_duplicates(subset="mun_norm").copy()
    s1_ref = s1_ref[["mun_name", "mun_norm", "plant_id", "prod_t", "dist_km"]].copy()
    s1_ref = s1_ref.rename(columns={
        "plant_id": "baseline_plant_id",
        "prod_t": "baseline_prod_t",
        "dist_km": "baseline_dist_km",
    })

    flows2_obs = normalize_observed_flows_to_baseline(flows2_obs, flows1_full)

    obs_prod = (
        flows2_obs.groupby("mun_norm", as_index=False)["prod_t"]
        .sum()
        .rename(columns={"prod_t": "observed_prod_t"})
    )

    closure = s1_ref.merge(obs_prod, on="mun_norm", how="left")
    closure["observed_prod_t"] = closure["observed_prod_t"].fillna(0.0)
    closure["residual_prod_t"] = closure["baseline_prod_t"] - closure["observed_prod_t"]
    closure["residual_prod_t"] = closure["residual_prod_t"].clip(lower=0.0)

    residual_rows = closure[closure["residual_prod_t"] > 0].copy()
    residual_rows = pd.DataFrame({
        "mun_name": residual_rows["mun_name"],
        "mun_norm": residual_rows["mun_norm"],
        "plant_id": residual_rows["baseline_plant_id"].astype(int),
        "prod_t": residual_rows["residual_prod_t"],
        "dist_km": residual_rows["baseline_dist_km"],
        "cost_per_t": np.nan,
    })

    flows_completed = pd.concat(
        [flows2_obs.copy(), residual_rows],
        ignore_index=True,
    )

    # Use the full S1 proximity baseline as the theoretical reference
    vor_completed = vor1_full.copy()

    return flows_completed, vor_completed


# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════
def main():
    print("=" * 74)
    print("  THREE-SCENARIO COMPARATIVE ANALYSIS")
    print("  Extremadura CDW (Construction & Demolition Waste) Network")
    print("=" * 74)

    # ---- Load all scenarios ----
    flows1, vor1 = load_scenario1()
    flows2, vor2 = load_scenario2()
    flows3, vor3 = load_scenario3()

    # ---- Compute indicators: full network / observed subnetwork ----
    r1 = compute_indicators(flows1, vor1, "1. Administrative (A01) - full network", multi_flow=False)
    flows2_norm = normalize_observed_flows_to_baseline(flows2, flows1)
    r2 = compute_indicators(flows2_norm, vor2, "2. Real flows (regional) - observed audited subnetwork", multi_flow=True)
    r3 = compute_indicators(flows3, vor3, "3. Optimised (A05) - full network", multi_flow=False)

    # Audited-support comparison: same municipality universe as S2 observed
    matched_munis = (
        set(flows2_norm["mun_norm"].unique())
        & set(vor2["mun_norm"].unique())
        & set(flows1["mun_norm"].unique())
        & set(flows3["mun_norm"].unique())
    )
    print(f"\n  Matched municipality support across S1/S2/S3 and S2-theoretical: {len(matched_munis)}")

    flows1_obs, vor1_obs = restrict_to_municipal_subset(flows1, vor1, matched_munis)
    flows2_obs, vor2_obs = restrict_to_municipal_subset(flows2_norm, vor2, matched_munis)
    flows3_obs, vor3_obs = restrict_to_municipal_subset(flows3, vor3, matched_munis)

    r1_obs = compute_indicators(flows1_obs, vor1_obs, "1obs. Administrative on matched audited support", multi_flow=False)
    r2_obs = compute_indicators(flows2_obs, vor2_obs, "2obs. Real flows on matched audited support", multi_flow=True)
    r3_obs = compute_indicators(flows3_obs, vor3_obs, "3obs. Optimised on matched audited support", multi_flow=False)

    # Full-network conservative closure for S2
    flows2_closure_seed, _ = restrict_to_municipal_subset(flows2_norm, vor2, set(flows2_norm["mun_norm"].unique()) & set(flows1["mun_norm"].unique()))
    flows2_full, vor2_full = build_s2_conservative_closure(flows2_closure_seed, flows1, vor1)
    r2_full = compute_indicators(flows2_full, vor2_full, "2full. Real flows with conservative full-network closure", multi_flow=True)

    # Anonymised public S2 input: municipality->plant assignment with volumes RESCALED to
    # theoretical production (raw commercial tonnages are withheld; see data availability note).
    flows2_full[["mun_norm", "plant_id", "dist_km", "prod_t"]].to_csv(
        OUT_DIR / "s2_observed_flows_public.csv", index=False)
    print(f"  [public export] s2_observed_flows_public.csv "
          f"({len(flows2_full)} flows, total prod {flows2_full['prod_t'].sum():,.0f} t)")

    # ═══════════════════════════════════════════════════════════════════
    # COMPARISON TABLE
    # ═══════════════════════════════════════════════════════════════════
    print("\n" + "=" * 74)
    print("  COMPARATIVE TABLE")
    print("=" * 74)

    indicators = [
        ("SCOPE", "", "", ""),
        ("  N plants",            f"{r1['N_plants']}",
                                  f"{r2['N_plants']}",
                                  f"{r3['N_plants']}"),
        ("  N municipalities",    f"{r1['N_municipalities']}",
                                  f"{r2['N_municipalities']}",
                                  f"{r3['N_municipalities']}"),
        ("  Total prod (t/yr)",   f"{r1['Total_prod_t']:,.0f}",
                                  f"{r2['Total_prod_t']:,.0f}",
                                  f"{r3['Total_prod_t']:,.0f}"),
        ("", "", "", ""),
        ("PLANT-LEVEL (means)", "", "", ""),
        ("  C_i (EUR/t)",         f"{r1['mean_C_i']:.2f}",
                                  f"{r2['mean_C_i']:.2f}",
                                  f"{r3['mean_C_i']:.2f}"),
        ("  U_i",                 "N/A", "N/A", "N/A"),
        ("  D_i^90 (km)",         f"{r1['mean_D90_i']:.1f}",
                                  f"{r2['mean_D90_i']:.1f}",
                                  f"{r3['mean_D90_i']:.1f}"),
        ("  Prec_i",              f"{r1['mean_Prec_i']:.3f}",
                                  f"{r2['mean_Prec_i']:.3f}",
                                  f"{r3['mean_Prec_i']:.3f}"),
        ("  Rec_i",               f"{r1['mean_Rec_i']:.3f}",
                                  f"{r2['mean_Rec_i']:.3f}",
                                  f"{r3['mean_Rec_i']:.3f}"),
        ("  Leak_dg_i (EUR/t)",   f"{r1['mean_Leak_dg_i']:.2f}",
                                  f"{r2['mean_Leak_dg_i']:.2f}",
                                  f"{r3['mean_Leak_dg_i']:.2f}"),
        ("  IET_i mean (%)",      f"{r1['mean_IET_i']:.1f}",
                                  f"{r2['mean_IET_i']:.1f}",
                                  f"{r3['mean_IET_i']:.1f}"),
        ("  IET_i median (%)",    f"{r1['median_IET_i']:.1f}",
                                  f"{r2['median_IET_i']:.1f}",
                                  f"{r3['median_IET_i']:.1f}"),
        ("", "", "", ""),
        ("NETWORK-LEVEL", "", "", ""),
        ("  C_bar (EUR/t)",       f"{r1['C_bar']:.2f}",
                                  f"{r2['C_bar']:.2f}",
                                  f"{r3['C_bar']:.2f}"),
        ("  Gini(C_m)",           f"{r1['Gini']:.4f}",
                                  f"{r2['Gini']:.4f}",
                                  f"{r3['Gini']:.4f}"),
        ("  MAD_C (EUR/t)",       f"{r1['MAD_C']:.2f}",
                                  f"{r2['MAD_C']:.2f}",
                                  f"{r3['MAD_C']:.2f}"),
        ("  CVaR_0.9(D) (km)",    f"{r1['CVaR_0.9(D)']:.1f}",
                                  f"{r2['CVaR_0.9(D)']:.1f}",
                                  f"{r3['CVaR_0.9(D)']:.1f}"),
        ("  micro-F1",            f"{r1['micro-F1']:.4f}",
                                  f"{r2['micro-F1']:.4f}",
                                  f"{r3['micro-F1']:.4f}"),
        ("  IGET (%)",            f"{r1['IGET']:.1f}",
                                  f"{r2['IGET']:.1f}",
                                  f"{r3['IGET']:.1f}"),
        ("  CV(T_i)",             f"{r1['CV(T_i)']:.3f}",
                                  f"{r2['CV(T_i)']:.3f}",
                                  f"{r3['CV(T_i)']:.3f}"),
        ("", "", "", ""),
        ("SYSTEM TOTALS", "", "", ""),
        ("  System cost (EUR/yr)", f"{r1['Total_system_cost_EUR']:,.0f}",
                                   f"{r2['Total_system_cost_EUR']:,.0f}",
                                   f"{r3['Total_system_cost_EUR']:,.0f}"),
        ("  Total t-km",           f"{r1['Total_tkm']:,.0f}",
                                   f"{r2['Total_tkm']:,.0f}",
                                   f"{r3['Total_tkm']:,.0f}"),
        ("  Irrat. leak (EUR/yr)", f"{r1['Irrational_leak_cost_EUR']:,.0f}",
                                   f"{r2['Irrational_leak_cost_EUR']:,.0f}",
                                   f"{r3['Irrational_leak_cost_EUR']:,.0f}"),
        ("", "", "", ""),
        ("REDUNDANCY CHECK", "", "", ""),
        ("  IGCR = 10000/IGET",   f"{r1['IGCR']:.2f}",
                                   f"{r2['IGCR']:.2f}",
                                   f"{r3['IGCR']:.2f}"),
        ("  IGSD = IGCR - 100",   f"{r1['IGSD']:.2f}",
                                   f"{r2['IGSD']:.2f}",
                                   f"{r3['IGSD']:.2f}"),
    ]

    audited_indicators = [
        ("SCOPE", "", "", ""),
        ("  N plants",            f"{r1_obs['N_plants']}",
                                  f"{r2_obs['N_plants']}",
                                  f"{r3_obs['N_plants']}"),
        ("  N municipalities",    f"{r1_obs['N_municipalities']}",
                                  f"{r2_obs['N_municipalities']}",
                                  f"{r3_obs['N_municipalities']}"),
        ("  Total prod (t/yr)",   f"{r1_obs['Total_prod_t']:,.0f}",
                                  f"{r2_obs['Total_prod_t']:,.0f}",
                                  f"{r3_obs['Total_prod_t']:,.0f}"),
        ("", "", "", ""),
        ("NETWORK-LEVEL", "", "", ""),
        ("  C_bar (EUR/t)",       f"{r1_obs['C_bar']:.2f}",
                                  f"{r2_obs['C_bar']:.2f}",
                                  f"{r3_obs['C_bar']:.2f}"),
        ("  Gini(C_m)",           f"{r1_obs['Gini']:.4f}",
                                  f"{r2_obs['Gini']:.4f}",
                                  f"{r3_obs['Gini']:.4f}"),
        ("  MAD_C (EUR/t)",       f"{r1_obs['MAD_C']:.2f}",
                                  f"{r2_obs['MAD_C']:.2f}",
                                  f"{r3_obs['MAD_C']:.2f}"),
        ("  CVaR_0.9(D) (km)",    f"{r1_obs['CVaR_0.9(D)']:.1f}",
                                  f"{r2_obs['CVaR_0.9(D)']:.1f}",
                                  f"{r3_obs['CVaR_0.9(D)']:.1f}"),
        ("  micro-F1",            f"{r1_obs['micro-F1']:.4f}",
                                  f"{r2_obs['micro-F1']:.4f}",
                                  f"{r3_obs['micro-F1']:.4f}"),
        ("  IGET (%)",            f"{r1_obs['IGET']:.1f}",
                                  f"{r2_obs['IGET']:.1f}",
                                  f"{r3_obs['IGET']:.1f}"),
        ("  Irrat. leak (EUR/yr)", f"{r1_obs['Irrational_leak_cost_EUR']:,.0f}",
                                   f"{r2_obs['Irrational_leak_cost_EUR']:,.0f}",
                                   f"{r3_obs['Irrational_leak_cost_EUR']:,.0f}"),
    ]

    closure_indicators = [
        ("SCOPE", "", "", ""),
        ("  N plants",            f"{r1['N_plants']}",
                                  f"{r2_full['N_plants']}",
                                  f"{r3['N_plants']}"),
        ("  N municipalities",    f"{r1['N_municipalities']}",
                                  f"{r2_full['N_municipalities']}",
                                  f"{r3['N_municipalities']}"),
        ("  Total prod (t/yr)",   f"{r1['Total_prod_t']:,.0f}",
                                  f"{r2_full['Total_prod_t']:,.0f}",
                                  f"{r3['Total_prod_t']:,.0f}"),
        ("", "", "", ""),
        ("NETWORK-LEVEL", "", "", ""),
        ("  C_bar (EUR/t)",       f"{r1['C_bar']:.2f}",
                                  f"{r2_full['C_bar']:.2f}",
                                  f"{r3['C_bar']:.2f}"),
        ("  Gini(C_m)",           f"{r1['Gini']:.4f}",
                                  f"{r2_full['Gini']:.4f}",
                                  f"{r3['Gini']:.4f}"),
        ("  MAD_C (EUR/t)",       f"{r1['MAD_C']:.2f}",
                                  f"{r2_full['MAD_C']:.2f}",
                                  f"{r3['MAD_C']:.2f}"),
        ("  CVaR_0.9(D) (km)",    f"{r1['CVaR_0.9(D)']:.1f}",
                                  f"{r2_full['CVaR_0.9(D)']:.1f}",
                                  f"{r3['CVaR_0.9(D)']:.1f}"),
        ("  micro-F1",            f"{r1['micro-F1']:.4f}",
                                  f"{r2_full['micro-F1']:.4f}",
                                  f"{r3['micro-F1']:.4f}"),
        ("  IGET (%)",            f"{r1['IGET']:.1f}",
                                  f"{r2_full['IGET']:.1f}",
                                  f"{r3['IGET']:.1f}"),
        ("  Irrat. leak (EUR/yr)", f"{r1['Irrational_leak_cost_EUR']:,.0f}",
                                   f"{r2_full['Irrational_leak_cost_EUR']:,.0f}",
                                   f"{r3['Irrational_leak_cost_EUR']:,.0f}"),
    ]

    # Print formatted table
    hdr = f"{'Indicator':<28s} {'1-Admin':>14s} {'2-Real':>14s} {'3-Optimised':>14s}"
    sep = "-" * len(hdr)
    print(hdr)
    print(sep)
    for row in indicators:
        if row[0] == "":
            print()
        elif row[1] == "":
            print(f"{row[0]}")
        else:
            print(f"  {row[0]:<26s} {row[1]:>14s} {row[2]:>14s} {row[3]:>14s}")

    print("\n" + "=" * 74)
    print("  AUDITED-SUPPORT COMPARISON (S1obs / S2obs / S3obs)")
    print("=" * 74)
    hdr = f"{'Indicator':<28s} {'S1obs':>14s} {'S2obs':>14s} {'S3obs':>14s}"
    sep = "-" * len(hdr)
    print(hdr)
    print(sep)
    for row in audited_indicators:
        if row[0] == "":
            print()
        elif row[1] == "":
            print(f"{row[0]}")
        else:
            print(f"  {row[0]:<26s} {row[1]:>14s} {row[2]:>14s} {row[3]:>14s}")

    print("\n" + "=" * 74)
    print("  FULL-NETWORK COMPARISON WITH CONSERVATIVE S2 CLOSURE")
    print("=" * 74)
    hdr = f"{'Indicator':<28s} {'S1full':>14s} {'S2full':>14s} {'S3full':>14s}"
    sep = "-" * len(hdr)
    print(hdr)
    print(sep)
    for row in closure_indicators:
        if row[0] == "":
            print()
        elif row[1] == "":
            print(f"{row[0]}")
        else:
            print(f"  {row[0]:<26s} {row[1]:>14s} {row[2]:>14s} {row[3]:>14s}")

    # ═══════════════════════════════════════════════════════════════════
    # ECONOMIC LOSS / SAVINGS ANALYSIS
    # ═══════════════════════════════════════════════════════════════════
    print("\n" + "=" * 74)
    print("  ECONOMIC COMPARISON")
    print("=" * 74)

    # S2 vs S1: how much more does the real system cost vs administrative
    # NOTE: direct comparison is tricky because S2 covers only 23 plants.
    # We compare on cost-per-tonne basis (C_bar).
    print("\n  [S2 vs S1] Real flows vs Administrative baseline:")
    print(f"    C_bar  S1 = {r1['C_bar']:.2f} EUR/t")
    print(f"    C_bar  S2 = {r2['C_bar']:.2f} EUR/t")
    delta_c_21 = r2["C_bar"] - r1["C_bar"]
    print(f"    Delta C_bar (S2 - S1) = {delta_c_21:+.2f} EUR/t")
    print(f"    NOTE: S1 and S2 now use the same cost model:")
    print(f"          C = {RHO} EUR/t-km * d_km + fixed_cost(T)/T + {V} EUR/t")

    # S2 vs S3: how much could be saved by optimising
    # Again, different scopes, so use cost-per-tonne
    print(f"\n  [S2 vs S3] Real flows vs Optimised network:")
    print(f"    C_bar  S2 = {r2['C_bar']:.2f} EUR/t")
    print(f"    C_bar  S3 = {r3['C_bar']:.2f} EUR/t")
    delta_c_23 = r2["C_bar"] - r3["C_bar"]
    print(f"    Delta C_bar (S2 - S3) = {delta_c_23:+.2f} EUR/t")
    if r2["Total_prod_t"] > 0:
        annual_saving_23 = delta_c_23 * r2["Total_prod_t"]
        print(f"    Potential annual saving (S2 prod * delta): {annual_saving_23:,.0f} EUR/yr")

    # S1 vs S3
    print(f"\n  [S1 vs S3] Administrative vs Optimised:")
    print(f"    C_bar  S1 = {r1['C_bar']:.2f} EUR/t")
    print(f"    C_bar  S3 = {r3['C_bar']:.2f} EUR/t")
    delta_c_13 = r1["C_bar"] - r3["C_bar"]
    print(f"    Delta C_bar (S1 - S3) = {delta_c_13:+.2f} EUR/t")
    annual_saving_13 = delta_c_13 * r1["Total_prod_t"]
    print(f"    Annual saving (S1 prod * delta): {annual_saving_13:,.0f} EUR/yr")

    # Transport efficiency comparison
    print(f"\n  Transport efficiency (IGET):")
    print(f"    S1: {r1['IGET']:.1f}%  |  S2: {r2['IGET']:.1f}%  |  S3: {r3['IGET']:.1f}%")
    print(f"    (100% = perfectly matches the road-distance proximity baseline)")

    # ═══════════════════════════════════════════════════════════════════
    # SAVE TO CSV
    # ═══════════════════════════════════════════════════════════════════
    print(f"\n{'='*74}")
    print(f"  SAVING RESULTS")
    print(f"{'='*74}")

    csv_rows = []
    for row in indicators:
        if row[0] == "" and row[1] == "":
            continue
        csv_rows.append({
            "Indicator": row[0].strip(),
            "S1_Administrative": row[1] if row[1] else "",
            "S2_Real_flows": row[2] if row[2] else "",
            "S3_Optimised": row[3] if row[3] else "",
        })

    # Add economic comparison rows
    csv_rows.append({"Indicator": "", "S1_Administrative": "",
                     "S2_Real_flows": "", "S3_Optimised": ""})
    csv_rows.append({"Indicator": "ECONOMIC COMPARISON",
                     "S1_Administrative": "", "S2_Real_flows": "",
                     "S3_Optimised": ""})
    csv_rows.append({"Indicator": "Delta C_bar S2-S1 (EUR/t)",
                     "S1_Administrative": "", "S2_Real_flows": f"{delta_c_21:+.2f}",
                     "S3_Optimised": ""})
    csv_rows.append({"Indicator": "Delta C_bar S2-S3 (EUR/t)",
                     "S1_Administrative": "", "S2_Real_flows": f"{delta_c_23:+.2f}",
                     "S3_Optimised": ""})
    csv_rows.append({"Indicator": "Delta C_bar S1-S3 (EUR/t)",
                     "S1_Administrative": f"{delta_c_13:+.2f}", "S2_Real_flows": "",
                     "S3_Optimised": ""})
    csv_rows.append({"Indicator": "Annual saving S1->S3 (EUR/yr)",
                     "S1_Administrative": f"{annual_saving_13:,.0f}",
                     "S2_Real_flows": "", "S3_Optimised": ""})
    if r2["Total_prod_t"] > 0:
        csv_rows.append({"Indicator": "Potential saving S2->S3 (EUR/yr)",
                         "S1_Administrative": "",
                         "S2_Real_flows": f"{annual_saving_23:,.0f}",
                         "S3_Optimised": ""})

    df_out = pd.DataFrame(csv_rows)
    df_out.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")
    print(f"  Saved: {OUT_CSV}")

    # Save plant-level scenario tables for downstream figures/tables in A03.
    def _export_plant_df(result, scenario_label, path):
        plant_df = result["_plant_df"].copy().reset_index()
        plant_df["scenario"] = scenario_label
        plant_df.to_csv(path, index=False, encoding="utf-8-sig")
        print(f"  Saved: {path}")

    _export_plant_df(r1, "S1_full", OUT_PLANT_S1)
    _export_plant_df(r2_full, "S2_full", OUT_PLANT_S2)
    _export_plant_df(r3, "S3_full", OUT_PLANT_S3)

    # ---- Notes ----
    print(f"\n{'='*74}")
    print("  NOTES")
    print(f"{'='*74}")
    print("  - Scenario 2 covers only 23 of 46 plants (regional operational data).")
    print("    Indicators are computed on the subset of municipalities present in the")
    print("    real data, NOT on all 383 municipalities of Extremadura.")
    print("  - U_i (capacity utilisation) is N/A for all scenarios (no capacity data).")
    print("  - Cost model shared across scenarios:")
    print(f"    S1/S2/S3: C = {RHO} * d_km + fixed_cost(T)/T + {V} EUR/t")
    print("  - Theoretical baseline:")
    print("    S1/S3: nearest-plant assignment by road distance (datos_red_real_municipios.csv)")
    print("    S2:    Theoretical assignment from Excel files (DATOS TEORICOS sections)")
    print("  - Consolidation map applied to S1/S3 baseline cells:")
    print(f"    {CONSOLIDATION}")

    print(f"\nDone.")


if __name__ == "__main__":
    main()
